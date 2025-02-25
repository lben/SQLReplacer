import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import os
from collections import defaultdict, deque

def extract_sheet_range(range_str):
    """Extract sheet name and range from a range string like 'Sheet1!A1:C10'."""
    if '!' in range_str:
        sheet_name, cell_range = range_str.split('!')
        # Remove quotes if present
        sheet_name = sheet_name.strip("'").strip('"')
        return sheet_name, cell_range
    return None, range_str

def get_column_range(start_col, end_col):
    """Get all columns between start_col and end_col inclusive."""
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)
    return [get_column_letter(idx) for idx in range(start_idx, end_idx + 1)]

def parse_formula_dependencies(formula):
    """
    Parse Excel formula to extract column dependencies, including:
    - Direct cell references (A1, B2)
    - Column ranges (A:C)
    - VLOOKUP references
    
    Returns a tuple of (column_dependencies, vlookup_dependencies)
    """
    dependencies = set()
    vlookup_deps = []
    
    # Remove string literals as they might contain things that look like cell references
    formula_without_strings = re.sub(r'"[^"]*"', '""', formula)
    formula_without_strings = re.sub(r"'[^']*'", "''", formula_without_strings)
    
    # Find cell references like A1, B2, etc.
    cell_refs = re.findall(r'([A-Z]+)(\d+)', formula_without_strings)
    for col, row in cell_refs:
        dependencies.add(col)
    
    # Find column ranges like A:C
    column_ranges = re.findall(r'([A-Z]+):([A-Z]+)', formula_without_strings)
    for start_col, end_col in column_ranges:
        for col in get_column_range(start_col, end_col):
            dependencies.add(col)
    
    # Find cell ranges like A1:C10
    cell_ranges = re.findall(r'([A-Z]+\d+):([A-Z]+\d+)', formula_without_strings)
    for start_cell, end_cell in cell_ranges:
        start_col = re.findall(r'([A-Z]+)', start_cell)[0]
        end_col = re.findall(r'([A-Z]+)', end_cell)[0]
        for col in get_column_range(start_col, end_col):
            dependencies.add(col)
    
    # Process VLOOKUP references
    # Example: VLOOKUP(A2,'Sheet2'!A:C,3,FALSE)
    vlookup_pattern = r'VLOOKUP\s*\((.*?),(.*?),(\d+)'
    vlookups = re.findall(vlookup_pattern, formula_without_strings, re.IGNORECASE)
    
    for lookup_value, table_array, col_index in vlookups:
        # Extract lookup value column dependencies
        lookup_col_refs = re.findall(r'([A-Z]+)(\d+)', lookup_value)
        for col, row in lookup_col_refs:
            dependencies.add(col)
        
        # Extract sheet name and range from table_array
        sheet_name, cell_range = extract_sheet_range(table_array)
        
        if sheet_name and cell_range:
            # It's a reference to another sheet
            # Extract column range
            if ':' in cell_range:
                range_parts = cell_range.split(':')
                if len(range_parts) == 2:
                    start_range = re.findall(r'([A-Z]+)', range_parts[0])
                    end_range = re.findall(r'([A-Z]+)', range_parts[1])
                    
                    if start_range and end_range:
                        start_col = start_range[0]
                        end_col = end_range[0]
                        # Store the VLOOKUP dependency
                        vlookup_deps.append((sheet_name, start_col, end_col, col_index))
    
    # Process INDEX/MATCH combinations which are commonly used alternatives to VLOOKUP
    # Example: INDEX(Sheet2!A:C, MATCH(A2, Sheet2!A:A, 0), 3)
    index_pattern = r'INDEX\s*\((.*?),(.*?),(.*?)\)'
    index_matches = re.findall(index_pattern, formula_without_strings, re.IGNORECASE)
    
    for array, row_num, col_num in index_matches:
        # Extract sheet name and range from array
        sheet_name, cell_range = extract_sheet_range(array)
        
        if sheet_name and cell_range:
            # It's a reference to another sheet
            # Extract column range
            if ':' in cell_range:
                range_parts = cell_range.split(':')
                if len(range_parts) == 2:
                    start_range = re.findall(r'([A-Z]+)', range_parts[0])
                    end_range = re.findall(r'([A-Z]+)', range_parts[1])
                    
                    if start_range and end_range:
                        start_col = start_range[0]
                        end_col = end_range[0]
                        # Store as a VLOOKUP-like dependency
                        vlookup_deps.append((sheet_name, start_col, end_col, "INDEX"))
        
        # Add dependencies from row_num and col_num expressions
        row_col_refs = re.findall(r'([A-Z]+)(\d+)', row_num + col_num)
        for col, row in row_col_refs:
            dependencies.add(col)
    
    return dependencies, vlookup_deps

def load_excel_partial(file_path, sheet_name, header_row, formula_row):
    """
    Load only the specified rows from an Excel sheet.
    Returns a dictionary with column data.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    sheet = wb[sheet_name]
    
    column_data = {}
    
    # Read header row
    for cell in sheet[header_row]:
        col_letter = cell.column_letter
        column_data[col_letter] = {'header': cell.value, 'formula': None, 'value': None}
    
    # Read formula row
    for cell in sheet[formula_row]:
        col_letter = cell.column_letter
        if col_letter in column_data:
            if cell.data_type == 'f':
                column_data[col_letter]['formula'] = cell.formula
            else:
                column_data[col_letter]['formula'] = None
                column_data[col_letter]['value'] = cell.value
    
    wb.close()
    return column_data

def load_sheet_for_vlookup(file_path, sheet_name, header_row, formula_row):
    """
    Load an entire sheet for VLOOKUP reference.
    Returns a dictionary with column data.
    """
    wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    
    # Handle case where sheet might not exist
    if sheet_name not in wb.sheetnames:
        print(f"Warning: Sheet '{sheet_name}' not found in workbook.")
        wb.close()
        return {}
        
    sheet = wb[sheet_name]
    
    column_data = {}
    
    # Read all columns in the sheet
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        
        # Get header from header row
        header_cell = sheet.cell(row=header_row, column=col_idx)
        header_value = header_cell.value
        
        # Get formula from formula row
        formula_cell = sheet.cell(row=formula_row, column=col_idx)
        
        if formula_cell.data_type == 'f':
            formula_value = formula_cell.formula
            value = None
        else:
            formula_value = None
            value = formula_cell.value
        
        column_data[col_letter] = {'header': header_value, 'formula': formula_value, 'value': value}
    
    wb.close()
    return column_data

def analyze_excel_dependencies(file_path, sheet_name, start_column, header_row, formula_row):
    """
    Analyze Excel dependencies starting from a specific column.
    
    Returns:
        A tuple of (dependencies, column_info)
    """
    # Dictionary to store dependencies for each sheet
    sheet_dependencies = {}
    
    # Dictionary to store column information for each sheet
    sheet_columns = {}
    
    # Set to keep track of processed columns for each sheet
    processed = defaultdict(set)
    
    # Set to track sheets that have been loaded
    loaded_sheets = set()
    
    # Queue for BFS traversal
    queue = deque([(sheet_name, start_column)])
    
    # Load initial sheet data
    try:
        sheet_columns[sheet_name] = load_excel_partial(file_path, sheet_name, header_row, formula_row)
        loaded_sheets.add(sheet_name)
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
        return {}, {}
    except Exception as e:
        print(f"Error loading Excel file: {str(e)}")
        return {}, {}
    
    # Initialize dependencies for the starting sheet
    sheet_dependencies[sheet_name] = defaultdict(set)
    
    while queue:
        current_sheet, current_column = queue.popleft()
        
        # Skip if already processed
        if current_column in processed[current_sheet]:
            continue
        
        # Mark as processed
        processed[current_sheet].add(current_column)
        
        # Initialize dependencies for this sheet if not already done
        if current_sheet not in sheet_dependencies:
            sheet_dependencies[current_sheet] = defaultdict(set)
        
        # Get column info
        if current_sheet not in sheet_columns:
            # Load sheet data if not already loaded
            try:
                sheet_columns[current_sheet] = load_excel_partial(file_path, current_sheet, header_row, formula_row)
                loaded_sheets.add(current_sheet)
            except Exception as e:
                print(f"Error loading sheet '{current_sheet}': {str(e)}")
                continue
        
        # Skip if column not found in sheet
        if current_column not in sheet_columns[current_sheet]:
            continue
        
        # Get formula for current column
        formula = sheet_columns[current_sheet][current_column].get('formula', '')
        
        # Process columns with formulas or values
        if formula or 'value' in sheet_columns[current_sheet][current_column]:
            if formula:
                # Parse dependencies
                dependencies, vlookup_deps = parse_formula_dependencies(formula)
                
                # Add dependencies to current column
                sheet_dependencies[current_sheet][current_column].update(dependencies)
                
                # Add dependencies to queue for processing
                for dep_col in dependencies:
                    if dep_col not in processed[current_sheet]:
                        queue.append((current_sheet, dep_col))
                
                # Process VLOOKUP dependencies
                for ref_sheet, start_col, end_col, col_index in vlookup_deps:
                    # Load referenced sheet if not already loaded
                    if ref_sheet not in loaded_sheets:
                        try:
                            sheet_columns[ref_sheet] = load_sheet_for_vlookup(file_path, ref_sheet, header_row, formula_row)
                            loaded_sheets.add(ref_sheet)
                        except Exception as e:
                            print(f"Error loading referenced sheet '{ref_sheet}': {str(e)}")
                            continue
                    
                    # Add VLOOKUP reference to dependencies
                    if col_index != "INDEX":  # Regular VLOOKUP
                        try:
                            lookup_col = get_column_letter(int(col_index))
                            sheet_dependencies[current_sheet][current_column].add(f"{ref_sheet}!{lookup_col}")
                        except ValueError:
                            # Handle cases where col_index is not a valid integer
                            print(f"Warning: Invalid column index in VLOOKUP: {col_index}")
                            pass
                    
                    # Add all columns in the range to process
                    for col in get_column_range(start_col, end_col):
                        if col not in processed[ref_sheet]:
                            queue.append((ref_sheet, col))
    
    return sheet_dependencies, sheet_columns

def generate_html_dependency_tree(dependencies, column_info):
    """Generate an HTML dependency tree from the dependencies."""
    html = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Excel Formula Dependency Tree</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                margin: 20px;
                line-height: 1.5;
                color: #333;
            }
            
            h1, h2 {
                color: #2c3e50;
                margin-top: 20px;
            }
            
            .sheet {
                margin-bottom: 30px;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 15px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            
            details {
                margin-left: 20px;
                margin-bottom: 10px;
                border-left: 1px solid #ddd;
                padding-left: 10px;
            }
            
            summary {
                cursor: pointer;
                font-weight: bold;
                margin-bottom: 5px;
                color: #2980b9;
            }
            
            summary:hover {
                color: #3498db;
            }
            
            .formula {
                font-family: Consolas, Monaco, 'Andale Mono', monospace;
                background-color: #f8f8f8;
                padding: 8px;
                margin-top: 5px;
                border-radius: 3px;
                border-left: 3px solid #3498db;
                overflow-x: auto;
                white-space: pre-wrap;
                word-break: break-word;
            }
            
            .value {
                font-family: Consolas, Monaco, 'Andale Mono', monospace;
                background-color: #f8f8f8;
                padding: 8px;
                margin-top: 5px;
                border-radius: 3px;
                border-left: 3px solid #27ae60;
                overflow-x: auto;
            }
            
            .circular {
                color: #e74c3c;
                font-weight: bold;
            }
            
            .no-deps {
                color: #95a5a6;
                font-style: italic;
                margin-left: 10px;
            }
            
            .sheet-title {
                display: flex;
                align-items: center;
                margin-bottom: 10px;
            }
            
            .sheet-name {
                margin-right: 10px;
            }
            
            .legend {
                margin-top: 20px;
                padding: 10px;
                background-color: #f8f8f8;
                border-radius: 5px;
            }
            
            .column-header {
                color: #7f8c8d;
                font-size: 0.9em;
            }
        </style>
    </head>
    <body>
        <h1>Excel Formula Dependency Tree</h1>
        <div class="legend">
            <p><strong>How to use:</strong> Click on a column to expand/collapse its dependencies. Formula are shown below each column.</p>
            <p><strong>Legend:</strong></p>
            <ul>
                <li><span style="color: #2980b9; font-weight: bold;">Column references</span> - Click to expand/collapse</li>
                <li><span class="formula" style="display: inline; padding: 2px 5px;">Formulas</span> - The Excel formula for the column</li>
                <li><span class="value" style="display: inline; padding: 2px 5px;">Values</span> - The data value for the column</li>
                <li><span class="circular">Circular references</span> - Detected circular dependencies</li>
            </ul>
        </div>
    """
    
    def build_tree(sheet_name, column, path=None):
        """
        Recursively build the dependency tree.
        
        Args:
            sheet_name: Name of the sheet
            column: Column to process
            path: Path of visited columns (to detect circular references)
        
        Returns:
            HTML string for this branch of the tree
        """
        if path is None:
            path = []
        
        # Check for circular references
        current_path = f"{sheet_name}!{column}"
        if current_path in path:
            return f"<div class='circular'>Circular reference: {current_path}</div>"
        
        # Add current column to path
        new_path = path + [current_path]
        
        # Get column info
        header = None
        formula = None
        value = None
        if sheet_name in column_info and column in column_info[sheet_name]:
            header = column_info[sheet_name][column].get('header', '')
            formula = column_info[sheet_name][column].get('formula', '')
            value = column_info[sheet_name][column].get('value', None)
        
        # Create HTML for this node
        result = f"<details><summary>{sheet_name}!{column}"
        if header:
            result += f" <span class='column-header'>({header})</span>"
        result += "</summary>"
        
        if formula:
            result += f"<div class='formula'>{formula}</div>"
        elif value is not None:
            result += f"<div class='value'>Value: {value}</div>"
        
        # Add dependencies
        if sheet_name in dependencies and column in dependencies[sheet_name]:
            deps = dependencies[sheet_name][column]
            if deps:
                for dep in sorted(deps):
                    if "!" in dep:
                        # Reference to another sheet
                        ref_sheet, ref_col = dep.split("!")
                        result += build_tree(ref_sheet, ref_col, new_path)
                    else:
                        # Reference to the same sheet
                        result += build_tree(sheet_name, dep, new_path)
            else:
                result += "<div class='no-deps'>No dependencies</div>"
        else:
            result += "<div class='no-deps'>No dependencies</div>"
        
        result += "</details>"
        return result
    
    # Build tree for each sheet
    sorted_sheets = sorted(dependencies.keys())
    for sheet_name in sorted_sheets:
        html += f"""
        <div class='sheet'>
            <div class='sheet-title'>
                <h2 class='sheet-name'>Sheet: {sheet_name}</h2>
            </div>
        """
        
        # Find root columns (columns that aren't dependencies of any other column in the same sheet)
        all_deps = set()
        for col_deps in dependencies[sheet_name].values():
            for dep in col_deps:
                if "!" not in dep:  # Only consider dependencies within the same sheet
                    all_deps.add(dep)
        
        sheet_columns = set(dependencies[sheet_name].keys())
        root_columns = sheet_columns - all_deps
        
        # If no root columns found, just use the first column alphabetically or the starting column
        if not root_columns:
            if sheet_columns:
                root_columns = {min(sheet_columns)}
            else:
                html += "<p>No formula columns found in this sheet.</p>"
                html += "</div>"
                continue
        
        # Build tree for each root column
        for column in sorted(root_columns):
            html += build_tree(sheet_name, column)
        
        html += "</div>"
    
    html += """
    </body>
    </html>
    """
    
    return html

def main(file_path, sheet_name, start_column, header_row, formula_row, output_html_path=None):
    """
    Main function to analyze Excel dependencies and generate HTML.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to analyze
        start_column: Starting column (e.g., 'A')
        header_row: Row number for headers (1-based)
        formula_row: Row number for formulas (1-based)
        output_html_path: Path to save HTML output, if None will use input file name with .html extension
    
    Returns:
        Path to the generated HTML file
    """
    print(f"Analyzing Excel dependencies in {file_path}")
    print(f"Starting from sheet: {sheet_name}, column: {start_column}")
    print(f"Header row: {header_row}, Formula row: {formula_row}")
    
    try:
        # Analyze dependencies
        dependencies, column_info = analyze_excel_dependencies(
            file_path, sheet_name, start_column, header_row, formula_row
        )
        
        if not dependencies or not column_info:
            print("No dependencies found or error occurred during analysis.")
            return None
        
        # Generate HTML
        html_content = generate_html_dependency_tree(dependencies, column_info)
        
        # Determine output path
        if output_html_path is None:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_html_path = f"{base_name}_dependency_tree.html"
        
        # Save HTML to file
        with open(output_html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"Dependency tree saved to: {output_html_path}")
        return output_html_path
    
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Analyze Excel formula dependencies')
    parser.add_argument('file_path', help='Path to the Excel file')
    parser.add_argument('sheet_name', help='Name of the sheet to analyze')
    parser.add_argument('start_column', help='Starting column (e.g., A)')
    parser.add_argument('header_row', type=int, help='Row number for headers (1-based)')
    parser.add_argument('formula_row', type=int, help='Row number for formulas (1-based)')
    parser.add_argument('--output', help='Path to save HTML output')
    
    args = parser.parse_args()
    
    main(args.file_path, args.sheet_name, args.start_column, args.header_row, args.formula_row, args.output)