from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import re
from functools import lru_cache
from typing import Dict, Set, List, Optional, Tuple
from dataclasses import dataclass
import os

@dataclass
class FormulaNode:
    column_name: str
    formula: str
    dependencies: List['FormulaNode']
    sheet_name: str

class ExcelFormulaAnalyzer:
    def __init__(self, excel_file: str):
        """Initialize the analyzer with the Excel file path."""
        self.excel_file = excel_file
        self._workbook = None
        self._headers_cache = {}
        self._formulas_cache = {}
        
    @property
    @lru_cache(maxsize=1)
    def workbook(self):
        """Cached property to load workbook only once."""
        if not self._workbook:
            self._workbook = load_workbook(self.excel_file, data_only=False)
        return self._workbook

    def get_headers(self, sheet_name: str, header_row: int = 1) -> Dict[str, str]:
        """Get column headers mapping (column letter to header name)."""
        if (sheet_name, header_row) not in self._headers_cache:
            sheet = self.workbook[sheet_name]
            headers = {}
            for cell in sheet[header_row]:
                if cell.value:  # Skip empty cells
                    col_letter = get_column_letter(cell.column)
                    headers[col_letter] = str(cell.value).strip()
            self._headers_cache[(sheet_name, header_row)] = headers
        return self._headers_cache[(sheet_name, header_row)]

    def get_formulas(self, sheet_name: str, formula_row: int = 2) -> Dict[str, str]:
        """Get formulas mapping (column letter to formula)."""
        if (sheet_name, formula_row) not in self._formulas_cache:
            sheet = self.workbook[sheet_name]
            formulas = {}
            for cell in sheet[formula_row]:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    col_letter = get_column_letter(cell.column)
                    formulas[col_letter] = cell.value
            self._formulas_cache[(sheet_name, formula_row)] = formulas
        return self._formulas_cache[(sheet_name, formula_row)]

    def _parse_cell_reference(self, ref: str) -> Tuple[Optional[str], str]:
        """Parse a cell reference into (sheet_name, column_letter)."""
        sheet_match = re.match(r"'?([^'!]+)'?!", ref)
        if sheet_match:
            sheet_name = sheet_match.group(1)
            col_ref = ref[sheet_match.end():]
        else:
            sheet_name = None
            col_ref = ref

        # Extract column letter from the reference
        col_match = re.match(r'(\$?[A-Z]+)', col_ref)
        if col_match:
            return sheet_name, col_match.group(1).replace('$', '')
        return None, ''

    def _extract_column_references(self, formula: str) -> List[Tuple[Optional[str], str]]:
        """Extract all column references from a formula."""
        # Handle VLOOKUP separately
        vlookup_pattern = r'VLOOKUP\((.*?),\s*([^,]+),\s*(\d+),\s*(?:TRUE|FALSE)\)'
        refs = []
        
        # First, handle VLOOKUP references
        for match in re.finditer(vlookup_pattern, formula, re.IGNORECASE):
            lookup_value, table_ref, _ = match.groups()
            # Add the lookup value reference if it's a cell reference
            if re.match(r'[A-Z]+\d+', lookup_value):
                sheet_name, col = self._parse_cell_reference(lookup_value)
                if col:
                    refs.append((sheet_name, col))
            
            # Add the table reference
            sheet_name, col = self._parse_cell_reference(table_ref)
            if col:
                refs.append((sheet_name, col))

        # Then handle regular cell references
        cell_pattern = r'(?:\'[^\']+\'!)?[$]?[A-Z]+[$]?\d+'
        for ref in re.finditer(cell_pattern, formula):
            sheet_name, col = self._parse_cell_reference(ref.group())
            if col:
                refs.append((sheet_name, col))

        return refs

    def build_dependency_tree(self, sheet_name: str, result_column: str, 
                            header_row: int = 1, formula_row: int = 2) -> FormulaNode:
        """Build a dependency tree starting from the result column."""
        headers = self.get_headers(sheet_name, header_row)
        formulas = self.get_formulas(sheet_name, formula_row)
        processed_nodes = set()  # To prevent circular dependencies

        def build_node(curr_sheet: str, col: str, processed: Set[str]) -> Optional[FormulaNode]:
            if col not in formulas:
                return None
            
            node_key = f"{curr_sheet}:{col}"
            if node_key in processed:
                return None  # Prevent circular dependencies
            processed.add(node_key)

            formula = formulas[col]
            dependencies = []
            
            for dep_sheet, dep_col in self._extract_column_references(formula):
                dep_sheet = dep_sheet or curr_sheet
                if dep_col:
                    dep_node = build_node(dep_sheet, dep_col, processed.copy())
                    if dep_node:
                        dependencies.append(dep_node)

            # Replace column references with header names in the formula
            readable_formula = formula
            for col_letter, header in headers.items():
                readable_formula = readable_formula.replace(col_letter, header)

            return FormulaNode(
                column_name=headers.get(col, col),
                formula=readable_formula,
                dependencies=dependencies,
                sheet_name=curr_sheet
            )

        return build_node(sheet_name, result_column, processed_nodes)

    def print_tree(self, node: FormulaNode, prefix: str = "", is_last: bool = True):
        """Print the formula dependency tree in a tree-like format."""
        if not node:
            return

        # Calculate branch characters
        branch = "└── " if is_last else "├── "
        
        # Print current node
        print(f"{prefix}{branch}[{node.sheet_name}]{node.column_name}: {node.formula}")
        
        # Calculate new prefix for children
        new_prefix = prefix + ("    " if is_last else "│   ")
        
        # Print dependencies
        for i, dep in enumerate(node.dependencies):
            is_last_dep = i == len(node.dependencies) - 1
            self.print_tree(dep, new_prefix, is_last_dep)

def analyze_excel_formulas(
    excel_file: str,
    sheet_name: str,
    result_column: str,
    header_row: int = 1,
    formula_row: int = 2
):
    """Main function to analyze Excel formulas and print the dependency tree."""
    analyzer = ExcelFormulaAnalyzer(excel_file)
    tree = analyzer.build_dependency_tree(
        sheet_name=sheet_name,
        result_column=result_column,
        header_row=header_row,
        formula_row=formula_row
    )
    
    print(f"\nFormula Dependency Tree for {os.path.basename(excel_file)}")
    print(f"Sheet: {sheet_name}, Result Column: {result_column}")
    print("-" * 50)
    analyzer.print_tree(tree)

# Example usage:
if __name__ == "__main__":
    # Replace these values with your Excel file details
    excel_file = "example.xlsx"
    sheet_name = "Sheet1"
    result_column = "D"  # The column containing the final result
    
    analyze_excel_formulas(
        excel_file=excel_file,
        sheet_name=sheet_name,
        result_column=result_column,
        header_row=1,
        formula_row=2
    )
