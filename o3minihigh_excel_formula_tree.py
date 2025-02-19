import re
import openpyxl
import sys
import argparse
from openpyxl.utils import get_column_letter

def load_workbook_data(filename, header_row=1, formula_row=2):
    """
    Load the workbook and, for each sheet, cache only the header row and the formula row.
    Returns a dictionary keyed by sheet name with:
       - 'headers': mapping of column letter -> header value
       - 'formulas': mapping of column letter -> formula string (if cell value is a formula)
    """
    wb = openpyxl.load_workbook(filename, data_only=False, read_only=True)
    workbook_data = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = {}
        formulas = {}

        # Read header row using enumerate to get the column index.
        header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
        for idx, cell in enumerate(header_cells, start=1):
            col_letter = get_column_letter(idx)
            headers[col_letter] = cell.value

        # Read formula row using enumerate as well.
        formula_cells = list(ws.iter_rows(min_row=formula_row, max_row=formula_row))[0]
        for idx, cell in enumerate(formula_cells, start=1):
            col_letter = get_column_letter(idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formulas[col_letter] = cell.value

        workbook_data[sheet] = {
            'headers': headers,
            'formulas': formulas
        }
    return workbook_data

def parse_formula_references(formula):
    """
    Parse a formula string and return a list of cell references.
    Each reference is a tuple (sheet, col, row) where 'sheet' is None if no sheet is referenced.
    
    This regex looks for patterns like:
       - A2
       - Sheet2!B2
       - 'My Sheet'!C2
    """
    # Pattern breakdown:
    #   (?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?  => optional sheet name (possibly quoted) followed by !
    #   (?P<col>[A-Z]{1,3})                     => column letters (A to Z, up to 3 letters)
    #   (?P<row>\d+)                           => row number
    pattern = r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?(?P<col>[A-Z]{1,3})(?P<row>\d+)"
    refs = []
    for match in re.finditer(pattern, formula):
        sheet = match.group('sheet')
        if sheet:
            sheet = sheet.strip("'")
        col = match.group('col')
        row = int(match.group('row'))
        refs.append((sheet, col, row))
    return refs

def substitute_formula(formula, current_sheet, workbook_data):
    """
    Replace cell references in the formula string with the header names.
    For example, if column B has header "profit", then a reference "B2" becomes "profit".
    If a reference has an explicit sheet (e.g. Sheet2!A2), that sheet is used.
    """
    pattern = r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_]+)!)?(?P<col>[A-Z]{1,3})(?P<row>\d+)"
    def repl(match):
        sheet = match.group('sheet')
        if sheet:
            sheet = sheet.strip("'")
        else:
            sheet = current_sheet
        col = match.group('col')
        header = workbook_data.get(sheet, {}).get('headers', {}).get(col)
        return str(header) if header is not None else match.group(0)
    return re.sub(pattern, repl, formula)

def build_dependency_tree(sheet, col, workbook_data, formula_row, header_row, visited=None):
    """
    Recursively build a dependency tree starting from the cell in (sheet, col).
    We assume that each formula is written on the same row (formula_row) so that
    if cell B in row formula_row has a formula, then any reference like B? (any row)
    is taken as referring to that same definition.
    
    A simple visited set is used to avoid infinite recursion in case of circular dependencies.
    """
    if visited is None:
        visited = set()
    key = (sheet, col)
    if key in visited:
        return {
            "sheet": sheet,
            "col": col,
            "header": workbook_data[sheet]['headers'].get(col, col),
            "formula": None,
            "children": [],
            "cycle": True
        }
    visited.add(key)
    
    formula = workbook_data[sheet]['formulas'].get(col)
    node = {
        "sheet": sheet,
        "col": col,
        "header": workbook_data[sheet]['headers'].get(col, col),
        "formula": formula,
        "children": []
    }
    
    if formula:
        # Parse the formula for any cell references.
        refs = parse_formula_references(formula)
        for ref_sheet, ref_col, ref_row in refs:
            # In our simplified model we ignore the actual row number because we assume
            # that every row uses the same formula logic.
            target_sheet = ref_sheet if ref_sheet else sheet
            if target_sheet in workbook_data:
                child_node = build_dependency_tree(target_sheet, ref_col, workbook_data, formula_row, header_row, visited)
                node["children"].append(child_node)
    visited.remove(key)
    return node

def print_tree(node, workbook_data, indent="", is_last=True):
    """
    Recursively print the dependency tree using a tree-like (Explorer-like) format.
    """
    branch = "└── " if is_last else "├── "
    if node.get("formula"):
        # Substitute cell references with header names.
        friendly = substitute_formula(node["formula"], node["sheet"], workbook_data)
        text = f"{node['header']} ({node['sheet']}!{node['col']}) = {friendly}"
    else:
        text = f"{node['header']} ({node['sheet']}!{node['col']})"
        if node.get("cycle"):
            text += " [cycle]"
    print(indent + branch + text)
    new_indent = indent + ("    " if is_last else "│   ")
    child_count = len(node.get("children", []))
    for idx, child in enumerate(node.get("children", [])):
        print_tree(child, workbook_data, new_indent, idx == (child_count - 1))

def main(filename, result_header="Result", formula_row=2, header_row=1, result_sheet=None):
    """
    Main function that loads the Excel file, finds the starting column by header name,
    builds the dependency tree, and prints it.
    
    Parameters:
      - filename: the Excel file to load.
      - result_header: the header name of the result column to start scanning from.
      - formula_row: the row number where formulas are defined (default is 2).
      - header_row: the row number where headers are defined (default is 1).
      - result_sheet: the sheet name to start with (default: first sheet).
    """
    workbook_data = load_workbook_data(filename, header_row, formula_row)
    if result_sheet is None:
        # Use the first sheet if no specific sheet is provided.
        result_sheet = list(workbook_data.keys())[0]
    
    # Find the column (by letter) in the chosen sheet that has the header matching result_header.
    target_col = None
    for col, header in workbook_data[result_sheet]['headers'].items():
        if header == result_header:
            target_col = col
            break
    if not target_col:
        print(f"Result column with header '{result_header}' not found in sheet '{result_sheet}'")
        return

    tree = build_dependency_tree(result_sheet, target_col, workbook_data, formula_row, header_row)
    print("Dependency Tree:")
    print_tree(tree, workbook_data)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build and print an Excel formula dependency tree.")
    parser.add_argument("filename", help="Path to the Excel file.")
    parser.add_argument("--result-header", default="Result", help="Header name of the result column to start from (default: 'Result').")
    parser.add_argument("--formula-row", type=int, default=2, help="Row number where formulas are defined (default: 2).")
    parser.add_argument("--header-row", type=int, default=1, help="Row number where headers are defined (default: 1).")
    parser.add_argument("--result-sheet", help="Name of the sheet to start with (default: first sheet).")

    args = parser.parse_args()
    main(args.filename, args.result_header, args.formula_row, args.header_row, args.result_sheet)
